#!/usr/bin/env python3
"""Portable regression tests for evidence, arithmetic and template preservation."""
import json
import tempfile
import unittest
from copy import deepcopy
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile
from openpyxl import Workbook
from extract_inputs import extract
from generate_docx import validate, expression, build, display
from check_docx import check

ROOT=Path(__file__).resolve().parents[1]


def fixture():
    def b(text='建议核对输入数据。',kind='proposal',sources=None):return dict(text=text,kind=kind,sources=sources or [])
    scene=dict(name='验证场景',description=b(),background=[b()],solution=[b()],provider=b('提供方待确认','missing'),data=[b('交付及频率未提供','missing')],payer=b('建议面向业务机构'),beneficiaries=[b()],revenue=dict(billing=b(),subject=b(),period='一年',parameters=[dict(id='N',label='数量',value=None,unit='家',kind='missing',sources=[]),dict(id='P',label='年度单价',value=None,unit='元/家/年',kind='missing',sources=[])],formula='N * P',result_name='年度营业收入',unit='元/年',costs=b('成本未提供，利润待测算','missing')),cases=[])
    return dict(document_mode='review',sources=[dict(id='S1',file='fixture.xlsx',locator='sheet:字段/row:2',excerpt='测试证据')],industries=[dict(title='测试行业空间场景应用',scenes=[scene])],allow_assumptions=False,limits=[])


class Invariants(unittest.TestCase):
    def test_formal_design_has_no_editorial_prefix(self):
        block=dict(text='通过接口同步数据，形成运行看板。',kind='proposal',sources=[])
        self.assertEqual(display(block),block['text'])

    def test_illustrative_case_is_allowed_without_fake_evidence(self):
        data=fixture()
        data['industries'][0]['scenes'][0]['cases']=[dict(type='example',title='区域机构运行示例',detail=dict(text='某区域机构接入工单记录，通过看板跟踪运行情况。',kind='proposal',sources=[]))]
        validate(data)

    def test_missing_parameters_never_become_zero(self):
        self.assertIsNone(validate(fixture())[0]['result'])

    def test_exact_decimal_calculation(self):
        params={'N':{'value':'0.1'},'P':{'value':'0.2'}}
        self.assertEqual(expression('N + P',params),Decimal('0.3'))

    def test_rejects_injection_even_with_missing_values(self):
        with self.assertRaises(ValueError):expression("N + __import__('os').getpid()",{'N':{'value':None}})

    def test_unknown_variable_and_division_by_zero(self):
        with self.assertRaises(ValueError):expression('X + 1',{})
        with self.assertRaises(ValueError):expression('N / 0',{'N':{'value':1}})

    def test_assumptions_require_permission(self):
        data=fixture(); p=data['industries'][0]['scenes'][0]['revenue']['parameters'][0]
        p.update(value='3',kind='assumption')
        with self.assertRaises(ValueError):validate(data)
        data['allow_assumptions']=True
        self.assertIsNone(validate(data)[0]['result'])

    def test_unknown_and_missing_evidence(self):
        data=fixture(); b=data['industries'][0]['scenes'][0]['description']
        b.update(kind='fact',sources=[])
        with self.assertRaises(ValueError):validate(data)
        b['sources']=['INVALID']
        with self.assertRaises(ValueError):validate(data)

    def test_multi_industry_package_and_escaping(self):
        data=fixture()
        second=deepcopy(data['industries'][0]);second['title']='另一行业'
        second['scenes'][0]['name']='含 & < > 的场景'
        data['industries'].append(second)
        with tempfile.TemporaryDirectory() as d:
            folder=Path(d); output=folder/'result.docx'; content=folder/'content.json'
            content.write_text(json.dumps(data,ensure_ascii=False),encoding='utf-8')
            build(data,output,ROOT/'assets/template.docx')
            self.assertTrue(check(output,content,ROOT/'assets/template.docx')['passed'])
            with ZipFile(output) as z:
                xml=z.read('word/document.xml').decode()
                self.assertIn('&amp;',xml)
                self.assertNotIn('{{',xml)
            with self.assertRaises(ValueError):build(data,output,ROOT/'assets/template.docx')

    def test_merged_category_preserves_anchor(self):
        with tempfile.TemporaryDirectory() as d:
            folder=Path(d); path=folder/'catalog.xlsx'
            w=Workbook();s=w.active;s.title='字段'
            s.append(['类别','字段']);s.append(['报警','报警时间']);s.append([None,'位置']);s.merge_cells('A2:A3')
            hidden=w.create_sheet('隐藏');hidden.sheet_state='hidden';hidden['A1']='说明'
            w.save(path)
            evidence=extract(path,folder/'out')
            row=next(r for r in evidence['records'] if r['locator']=='sheet:字段/row:3')
            self.assertEqual(row['text'][0]['value'],'报警')
            self.assertEqual(row['text'][0]['merged_anchor'],'A2')
            self.assertEqual(evidence['sheets'][1]['state'],'hidden')


if __name__=='__main__':unittest.main(verbosity=2)
